import asyncio
import openpyxl
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright


URL = "https://infosec-conferences.com/"


async def safe_inner_text(element):
    try:
        return (await element.inner_text()).strip()
    except Exception:
        return ""


async def get_event_detail(page, url):
    try:
        await page.goto(url, timeout=60000, wait_until="domcontentloaded")
        await asyncio.sleep(1)  # pausa para asegurar carga JS

        # Intentamos extraer descripción o contenido principal (puede variar según evento)
        desc = ""
        desc_selector_candidates = [
            "div.event-description",
            "div#eventContent",
            "section.description",
            "article",
            "div.content"
        ]
        for selector in desc_selector_candidates:
            el = await page.query_selector(selector)
            if el:
                desc = await safe_inner_text(el)
                if desc:
                    break
        return desc
    except Exception as e:
        return f"Error accediendo detalle: {e}"


async def scrape_all(output_xlsx="infosec_conferences_full.xlsx"):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)  # headless=False para debug visual
        context = await browser.new_context()
        main_page = await context.new_page()

        try:
            await main_page.goto(URL, timeout=60000, wait_until="domcontentloaded")
            print("Página cargada correctamente")
        except Exception as e:
            print(f"Error cargando la página: {e}")
            await browser.close()
            return

        await main_page.wait_for_selector("table.table-striped tbody tr", timeout=60000)

        # Crear workbook Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Eventos Ciberseguridad"

        # Extraer encabezados de tabla y añadir columna detalle
        header_cells = await main_page.query_selector_all("table.table-striped thead tr th")
        headers = [await safe_inner_text(cell) for cell in header_cells]
        headers.append("Descripción Detallada")
        ws.append(headers)

        total_events = 0
        total_pages = 0

        # Página para detalles
        detail_page = await context.new_page()

        while True:
            rows = await main_page.query_selector_all("table.table-striped tbody tr")
            for row in rows:
                cols = await row.query_selector_all("td")
                row_data = []
                event_url = None
                for idx, col in enumerate(cols):
                    text = await safe_inner_text(col)
                    row_data.append(text)
                    # Capturar URL del evento desde la primera columna (normalmente es link)
                    if idx == 0:
                        link = await col.query_selector("a")
                        if link:
                            event_url = await link.get_attribute("href")

                if event_url and event_url.startswith("/"):
                    event_url = URL.rstrip("/") + event_url

                description = ""
                if event_url:
                    description = await get_event_detail(detail_page, event_url)
                    await asyncio.sleep(0.5)  # pausa para no saturar servidor

                row_data.append(description)
                ws.append(row_data)

            total_events += len(rows)
            total_pages += 1
            print(f"Página {total_pages} scrapeada, eventos en esta página: {len(rows)}")

            next_button = await main_page.query_selector("ul.pagination li.page-item.next:not(.disabled) a")
            if next_button:
                await next_button.click()
                await main_page.wait_for_load_state('networkidle')
                await main_page.wait_for_selector("table.table-striped tbody tr", timeout=60000)
                await asyncio.sleep(1)  # pausa para estabilizar
            else:
                break

        # Ajustar ancho columnas en Excel
        for col_num, column_title in enumerate(headers, 1):
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[get_column_letter(col_num)]
            )
            adjusted_width = max_length + 5
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        wb.save(output_xlsx)

        await browser.close()
        print(f"Scraped {total_events} eventos en {total_pages} páginas. Datos guardados en '{output_xlsx}'.")


if __name__ == "__main__":
    asyncio.run(scrape_all())
